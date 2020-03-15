import xlrd  # 读一般使用这个
import openpyxl
import json

f = open("admins.json", "r", encoding="utf-8")
admin_dic = json.load(f)
f.close()
sn_admin_dic = admin_dic["data"]

workbook = xlrd.open_workbook(r'./_1_hosts_list.xlsx')
# 获取所有sheet
sheet_name = workbook.sheet_names()[1]
# print(sheet_name)

# 根据sheet索引或者名称获取sheet内容
sheet = workbook.sheet_by_name('Sheet2')
# print(sheet.cell_value(0,1))
# print (sheet.name,sheet.nrows,sheet.ncols)
index_name = sheet.row_values(0) # 获取第1行内容
# print(index_name)
new_index = ['机群ID', '机群名称', '机器类别（vps虚机；srvs实机）', 'IP', 'SN']
#index_name = ['ID', 'PROJECT_CODE', 'PROJECT_NAME', 'TIME', 'TIME_UNIT', 'UPDATE_TIME', '机群ID', '机群名称', '机器类别（vps虚机；srvs实机）', 'DOMAIN', 'IP', 'SN', 'CPU_IDLE_RATE', 'MEM_IDLE_RATE', 'DISK_IDLE_RATE', 'GPU_IDLE_RATE', 'SERVER_IDLE_RATE']
# for i in new_index:
#     print(i, index_name.index(i))

fp = openpyxl.Workbook() #创建工作簿
sheet1 = fp.create_sheet()

new_index_name = ['SN', '机群ID', '机群名称', '机器类别（vps虚机；srvs实机）', 'IP', '机器管理员']
for i in range(len(new_index_name)):
    sheet1.cell(row=1, column=i+1).value = new_index_name[i]

mail_set = set()
for i in range(1, sheet.nrows):
    row = sheet.row_values(i)
    # sn号
    sn = row[11]
    # 集群ID
    cluster_id = row[6]
    # 机群名称
    cluster_name = row[7]

    # 机器类别
    host_class = row[8]
    if host_class.strip() == 'vps':
        h_class = "虚机"
    elif host_class.strip() == 'srv':
        h_class = "实机"
    else:
        h_class = host_class
    # ip
    ip = row[10]
    # 机器管理员
    ip_admins = sn_admin_dic[sn]
    host_admins = ",".join(ip_admins)
    mail_set.add(host_admins)

    sheet1_row = i+1
    sheet1.cell(row=sheet1_row, column=1).value = sn
    sheet1.cell(row=sheet1_row, column=2).value = cluster_id
    sheet1.cell(row=sheet1_row, column=3).value = cluster_name
    sheet1.cell(row=sheet1_row, column=4).value = h_class
    sheet1.cell(row=sheet1_row, column=5).value = ip
    sheet1.cell(row=sheet1_row, column=6).value = host_admins
fp.save("newhost_list.xlsx")

person_mail = set()
for i in mail_set:
    li = i.split(",")
    for j in li:
        person_mail.add("{}@sogou-inc.com".format(j.strip()))
print(";".join(list(person_mail)))
print("done!")
