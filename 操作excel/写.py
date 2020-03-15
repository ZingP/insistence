import sys
import json
import openpyxl
# 通常写excel使用下面这个模块
import xlsxwriter

def load_data(data_file):
    """加载数据"""
    data_list = []
    with open(data_file, "r") as f:
        for line in f:
            data_dic = json.loads(line)
            data_list.append(data_dic)
    return data_list


def list2excel(data_list, file_path):
    fp = openpyxl.Workbook() #创建工作簿
    sheet1 = fp.create_sheet()
    column_names = ["中心词", "节目id", "节目名称","节目摘要"]
    # 将列名写入第一行
    for i in range(len(column_names)):
        sheet1.cell(row=1, column=i+1).value = column_names[i]

    # 将数据写入
    for i, item in enumerate(data_list):
        
        center_word = item["center_word"]
        audio_id  = item["audio_id"]
        audio_name = item["audio_name"]
        audio_info = item["audio_info"]

        sheet1_row = i+2
        sheet1.cell(row=sheet1_row, column=1).value = center_word
        sheet1.cell(row=sheet1_row, column=2).value = audio_id
        sheet1.cell(row=sheet1_row, column=3).value = audio_name
        sheet1.cell(row=sheet1_row, column=4).value = audio_info
    fp.save(file_path)


def write2excel(ata_list, file_path):
    """写Excel"""
    workbook = xlsxwriter.Workbook(file_path)
    sheet1 = workbook.add_worksheet()
    column_names = ["中心词", "节目id", "节目名称","节目摘要"]
    for i in range(len(column_names)):
        sheet1.write(0, i, column_names[i])
        # 将数据写入
    for i, item in enumerate(data_list):
        
        center_word = item["center_word"]
        audio_id  = item["audio_id"]
        audio_name = item["audio_name"]
        audio_info = item["audio_info"]

        sheet1_row = i+1
        sheet1.write(sheet1_row, 0, center_word)
        sheet1.write(sheet1_row, 1, audio_id)
        sheet1.write(sheet1_row, 2, audio_name)
        sheet1.write(sheet1_row, 3, audio_info)
    workbook.close()

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Useage: python3 json2excel.py filename")
    opt = sys.argv[1:]
    data_file = opt[0]
    excel_file = "{}.xlsx".format(data_file)
    data_list = load_data(data_file)
    write2excel(data_list, excel_file)
    print("Translating {} to {} done!".format(data_file, excel_file))
